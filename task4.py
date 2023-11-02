import os
import openpyxl
import time
from datetime import datetime

class BusRecorder:
    number_of_student_on_bus = 0
    def __init__(self, _ID, _task_Ready, _task_Data):
        print("Init task 4")

        self.number_of_seats = 3

        self.folder_path = ".\\excel\\buslog\\" #the path to the folder that stores data
        current_date = datetime.now().strftime("%Y-%m-%d") #the day 
        self.excel_file_path = os.path.join(self.folder_path, f"{current_date}.xlsx") #the path to the file excel of current day
        
        #if the file excel of current day is not exists, it will create ones_ this can ensure that one day only has one 1 file Excel
        if not os.path.exists(self.excel_file_path):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "buslog"
            row_to_insert = ["Student ID", "Full name", "Time"]
            self.workbook['buslog'].append(row_to_insert)
            self.workbook.save(self.excel_file_path)

        self.ID = _ID
        self.task_Ready = _task_Ready
        self.task_Data = _task_Data

    def Run(self):
        '''Return "Full" if the bus is full, "Available" if the bus still has empty seat(s).'''
        
        if (not self.task_Ready[self.ID]):
            return "[i]"

        student_id, student_name = self.task_Data[self.ID].split('|')

        print("#"*30)
        print(f"Adding {student_name} to the bus log file...")

        #update information to file Excel
        
        if (self.number_of_student_on_bus < self.number_of_seats):
            self.workbook = openpyxl.load_workbook(self.excel_file_path)
            self.sheet = self.workbook['buslog']
            
            # Check if the student have already entered the bus
            update = False
            for row in self.sheet.iter_rows(min_row=2, values_only=True):  #Start form the second row, because the first one is title
                if row[0] == student_id:
                    update = True
                    break

            # Only update bus log if that student haven't entered the bus beforehand
            if not update:
                current_time = time.strftime("%H:%M:%S %d-%m-%Y")
                
                row_to_insert = [student_id, student_name, current_time]
                
                self.sheet.append(row_to_insert)
                self.workbook.save(self.excel_file_path)
                self.number_of_student_on_bus +=1

        else:
            print("Already full\n")


        
        # Last task in the chain
        self.task_Ready[self.ID] = False
        # Restart the cycle: return to the first task
        self.task_Ready[0] = True

        bus_state = str(self.number_of_student_on_bus) + '/' + str(self.number_of_seats)

        return "[3]" + bus_state

if __name__ == "__main__":
    bus = BusRecorder()
    print(bus.Run("10422117", "ABC"))

    print(bus.Run("10422075", "DEF"))

    print(bus.Run("10421101", "GHK"))

    print(bus.Run("10422052", "WWW"))

    print(bus.Run("10422118", "ZZZ"))
